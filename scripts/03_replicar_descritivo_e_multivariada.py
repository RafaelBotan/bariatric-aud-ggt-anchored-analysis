"""
03 - Replicar análise descritiva original e rodar multivariada
==============================================================
Objetivos:
  A) Confirmar que reproduzimos os números do docx (sanity check)
  B) Rodar regressão logística multivariada (NÃO foi feita pela Mariana)
  C) Aplicar correção FDR Benjamini-Hochberg
  D) Excluir variáveis tautológicas (bebida atual/anterior/tipo) do modelo principal

Definições conforme docx + dicionário:
  CAGE_PONTOS: 0 = baixo risco; ≥1 = algum grau (=> 'leve/intenso' usa ≥1)
    docx: leve/intenso = 26+8 = 34 ou 26 (algum grau) + 8 (provável)
    Na docx Tabela 2: leve/intenso n=34, baixo n=96 → cutoff CAGE>=1
  AUDIT_PONTOS:
    docx Tabela 2: dependência n=32 (4.6+4.6+15.4=24.6%), baixo n=98
    Categorias: 0-7 baixo; 8-15 risco; 16-19 nocivo; 20+ dependência
    docx classifica "dependência" como 8+ (consumo risco + uso nocivo + provável dependência)
    Confirmar: 32 = 6 + 6 + 20 = AUDIT >= 8
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd, numpy as np
import openpyxl
from scipy import stats

# Carrega planilha original (139)
wb = openpyxl.load_workbook('Y:/Estudos_Arruda_Galvao/Estudo Alcool Arruda/Planilha - 140 pacientes (1) (8).xlsx',
                            data_only=True)
ws = wb['Planilha1']
HDR = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    nome = ws.cell(r, 1).value
    if not nome:
        continue
    row = {'NOME': str(nome).strip()}
    for c, h in enumerate(HDR, 1):
        if h:
            row[h] = ws.cell(r, c).value
    rows.append(row)
df = pd.DataFrame(rows)
print(f'Planilha: {len(df)}')

# Numerizar
for c in ['CAGE_PONTOS','AUDIT_PONTOS','IDADE','Tempo_cir','IMC_pre','IMC_Min','IMC_atual',
          'Sexo:','Estado Civil','Escolaridade','religião','ETNIA','RENDA','violencia','filhos',
          'pai_bebida','mae_bebida','Bebida atua','Bebida ant','tipo','preju amiza','preju famili',
          'preju finan','preju trab']:
    df[c] = pd.to_numeric(df[c], errors='coerce')

# === Restringir a 130 (linkados) ===
cw = pd.read_csv('Y:/Estudos_Arruda_Galvao/Estudo Alcool Arruda/trabalho/dados/CROSSWALK_ALCOOL_ARRUDA.csv',
                 low_memory=False)
linkados_nomes = set(cw['NOME_ALC'])
df_link = df[df['NOME'].isin(linkados_nomes)].copy()
print(f'Linkados: {len(df_link)}')

# Ainda assim, vou rodar nos 139 e 130 separadamente para comparar com docx
def desfechos(d):
    d = d.copy()
    d['CAGE_pos'] = (d['CAGE_PONTOS'] >= 1).astype(int)
    d['AUDIT_pos'] = (d['AUDIT_PONTOS'] >= 8).astype(int)
    return d

df = desfechos(df)
df_link = desfechos(df_link)

print('\n=== Tabela 2 docx vs replicação ===')
print('docx N=130; CAGE+ n=34, AUDIT+ n=32')
print(f'replic 139: CAGE+ n={df["CAGE_pos"].sum()} ({100*df["CAGE_pos"].mean():.1f}%); AUDIT+ n={df["AUDIT_pos"].sum()} ({100*df["AUDIT_pos"].mean():.1f}%)')
print(f'replic 130: CAGE+ n={df_link["CAGE_pos"].sum()} ({100*df_link["CAGE_pos"].mean():.1f}%); AUDIT+ n={df_link["AUDIT_pos"].sum()} ({100*df_link["AUDIT_pos"].mean():.1f}%)')

# Vou usar os 130 da análise oficial = nossos 130 linkados
# Salvar este dataset trabalhado como o oficial
df_link.to_csv('Y:/Estudos_Arruda_Galvao/Estudo Alcool Arruda/trabalho/dados/DATASET_130.csv',
               index=False, encoding='utf-8')
print('\nDataset salvo em DATASET_130.csv')

# === Tabela 1 ===
print('\n=== Tabela 1 docx (descritivo numéricas) ===')
for var in ['IDADE','Tempo_cir','IMC_pre','IMC_Min','IMC_atual','CAGE_PONTOS','AUDIT_PONTOS']:
    s = df_link[var].dropna()
    print(f'  {var}: n={len(s)} min={s.min():.1f} max={s.max():.1f} média={s.mean():.1f}±{s.std():.1f} mediana={s.median():.1f} [P25={s.quantile(0.25):.1f}-P75={s.quantile(0.75):.1f}]')

# === Multivariada Logística ===
# Variáveis preditoras (NÃO-tautológicas):
#   sexo, idade, tempo_cir, escolaridade, religião (sem religião=4 dummy), etnia, renda, filhos
#   pai_bebida, mae_bebida, violência geral, prejuízos (familiar, financeiro, amizade, trabalho)
# EXCLUÍDAS (tautológicas com AUDIT/CAGE):
#   bebida atual, bebida anterior, tipo de bebida — porque definem o consumo.
import statsmodels.api as sm
from statsmodels.stats.multitest import multipletests

print('\n========================================')
print('MULTIVARIADA — desfecho AUDIT >= 8')
print('========================================')

d = df_link.copy()
d['masc'] = (d['Sexo:'] == 1).astype(int)
d['idade_anos'] = d['IDADE']
d['tempo_cir_anos'] = d['Tempo_cir']/12.0
d['imc_atual'] = d['IMC_atual']
d['pai_bebia'] = (d['pai_bebida'] == 1).astype(int)
d['mae_bebia'] = (d['mae_bebida'] == 1).astype(int)
d['violencia_qq'] = (d['violencia'] == 1).astype(int)
d['prej_amiza'] = (d['preju amiza'] == 1).astype(int)
d['prej_famili'] = (d['preju famili'] == 1).astype(int)
d['prej_finan'] = (d['preju finan'] == 1).astype(int)
d['prej_trab'] = (d['preju trab'] == 1).astype(int)
d['ensino_sup'] = (d['Escolaridade'] >= 4).astype(int)  # 4 ou 5 = superior
d['sem_religiao'] = (d['religião'] == 2).astype(int)  # 2=sem religião
d['filhos_sim'] = (d['filhos'] == 1).astype(int)

# Modelo principal sem confundidores tautológicos
X_cols = ['masc','idade_anos','tempo_cir_anos','ensino_sup','sem_religiao',
          'pai_bebia','violencia_qq','prej_famili']
def ajusta(d, y, X_cols, label):
    sub = d[[y]+X_cols].dropna()
    print(f'\n--- {label} (n={len(sub)} de {len(d)}) ---')
    X = sm.add_constant(sub[X_cols])
    try:
        m = sm.Logit(sub[y], X).fit(disp=0, maxiter=100)
        params = m.params
        ci = m.conf_int()
        pv = m.pvalues
        out = pd.DataFrame({
            'OR': np.exp(params),
            'IC95_low': np.exp(ci[0]),
            'IC95_high': np.exp(ci[1]),
            'p_value': pv,
        })
        print(out.to_string())
        # FDR
        pvs = pv.drop('const')
        rej, p_fdr, _, _ = multipletests(pvs, method='fdr_bh')
        print('\nFDR-BH:')
        for v, p, q in zip(pvs.index, pvs.values, p_fdr):
            print(f'  {v}: p={p:.3f}  q={q:.3f}  {"sig" if q<0.05 else ""}')
        return out
    except Exception as e:
        print(f'ERRO: {e}')
        return None

out_audit = ajusta(d, 'AUDIT_pos', X_cols, 'AUDIT >= 8')
out_cage  = ajusta(d, 'CAGE_pos', X_cols, 'CAGE >= 1')

# Modelo só com preditores demográficos (mais parcimonioso)
print('\n=== MODELO PARCIMONIOSO (apenas demográfico+familiar) ===')
ajusta(d, 'AUDIT_pos', ['masc','idade_anos','tempo_cir_anos','pai_bebia'], 'AUDIT (parcim.)')
ajusta(d, 'CAGE_pos', ['masc','idade_anos','tempo_cir_anos','pai_bebia'], 'CAGE (parcim.)')

# Salvar
if out_audit is not None:
    out_audit.to_csv('Y:/Estudos_Arruda_Galvao/Estudo Alcool Arruda/trabalho/resultados/03_multivariada_AUDIT.csv')
if out_cage is not None:
    out_cage.to_csv('Y:/Estudos_Arruda_Galvao/Estudo Alcool Arruda/trabalho/resultados/03_multivariada_CAGE.csv')
print('\nSalvo.')
